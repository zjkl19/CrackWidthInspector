from __future__ import annotations

import argparse
import csv
import json
import math
import statistics
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_SUMMARY = Path("outputs/platform_unmarked_full_20260527_v2/combined_summary.csv")
DEFAULT_POINTS_CSV = Path("outputs/platform_unmarked_full_20260527_v2/manual_point_review_points.csv")
DEFAULT_POINTS_XLSX = Path("outputs/platform_unmarked_full_20260527_v2/manual_point_review_points.xlsx")
DEFAULT_EVAL_DIR = Path("outputs/platform_unmarked_full_20260527_v2/manual_point_review_eval")
DEFAULT_POINT_CONFIG_CSV = Path("outputs/platform_unmarked_full_20260527_v2/sensor_point_config.csv")

MANUAL_WIDTH_DIGITS = 2
MANUAL_PREFILL_SOURCE = "prefilled_auto_2dp"
MANUAL_CONFIRMED_SOURCE = "manual_confirmed"
MANUAL_BATCH_CURRENT_IMAGE_SOURCE = "batch_confirmed_current_image"
MANUAL_BATCH_SELECTED_IMAGES_SOURCE = "batch_confirmed_selected_images"
MANUAL_BATCH_ALL_IMAGES_SOURCE = "batch_confirmed_all_images"
MANUAL_ACTUAL_OVERRIDE_SOURCE = "manual_actual_override"
MANUAL_INSTRUMENT_NAME = "裂缝宽度观测仪"
MANUAL_INSTRUMENT_MODEL = "HC-CK101"
MANUAL_INSTRUMENT_MANUFACTURER = "北京海创高科技有限公司"
MANUAL_INSTRUMENT_RANGE = "0~3 mm"
MANUAL_INSTRUMENT_UNCERTAINTY = "U=0.01 mm, k=2"
MANUAL_INSTRUMENT_BASIS = "JJF 1334-2012 混凝土裂缝宽度及深度测量仪校准规范"

POINT_FIELDS = [
    "point_id",
    "image_record_id",
    "point_order",
    "target_position_pct",
    "actual_position_pct",
    "sn",
    "device_name",
    "uptime",
    "filename",
    "input_image_path",
    "overlay_path",
    "profile_csv_path",
    "target_x_px",
    "target_y_px",
    "target_search_radius_px",
    "x_px",
    "y_px",
    "left_x_px",
    "left_y_px",
    "right_x_px",
    "right_y_px",
    "auto_width_mm",
    "auto_width_px",
    "mask_width_mm",
    "contrast",
    "threshold",
    "quality_usable",
    "marker_contaminated",
    "selection_reason",
    "replacement_from_target_px",
    "point_config_source",
    "actual_override_source",
    "actual_override_distance_px",
    "original_actual_position_pct",
    "original_x_px",
    "original_y_px",
    "original_left_x_px",
    "original_left_y_px",
    "original_right_x_px",
    "original_right_y_px",
    "original_auto_width_mm",
    "original_auto_width_px",
    "original_mask_width_mm",
    "original_contrast",
    "original_threshold",
    "original_selection_reason",
    "review_usable",
    "exclude_reason",
    "image_deleted",
    "delete_reason",
    "manual_width_mm",
    "manual_source",
    "review_status",
    "manual_note",
]

POINT_CONFIG_FIELDS = [
    "sn",
    "device_name",
    "point_order",
    "point_name",
    "target_x_px",
    "target_y_px",
    "search_radius_px",
    "enabled",
    "note",
    "updated_from_image",
]


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return None if math.isnan(number) else number
    text = str(value).strip()
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return None if math.isnan(number) else number


def is_truthy(value: Any) -> bool:
    text = str(value if value is not None else "").strip().lower()
    return text in {"1", "true", "yes", "y", "是", "已删除", "deleted"}


def format_manual_width(value: Any) -> str:
    number = to_float(value)
    if number is None:
        return ""
    return f"{number:.{MANUAL_WIDTH_DIGITS}f}"


def default_manual_width(row: dict[str, Any]) -> str:
    return format_manual_width(row.get("auto_width_mm"))


def normalize_point_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(row)
    for field in POINT_FIELDS:
        normalized.setdefault(field, "")
    status = str(normalized.get("review_status") or "").strip().lower()
    if not status:
        status = "pending"
    normalized["review_status"] = status
    normalized["image_deleted"] = "1" if is_truthy(normalized.get("image_deleted")) else "0"
    normalized["delete_reason"] = str(normalized.get("delete_reason") or "")

    manual = format_manual_width(normalized.get("manual_width_mm"))
    source = str(normalized.get("manual_source") or "").strip()
    if not manual:
        manual = default_manual_width(normalized)
        if manual and not source:
            source = MANUAL_PREFILL_SOURCE
    elif not source:
        source = MANUAL_CONFIRMED_SOURCE if status in {"reviewed", "confirmed"} else "manual_existing"
    normalized["manual_width_mm"] = manual
    normalized["manual_source"] = source
    return normalized


def normalize_point_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [normalize_point_row(row) for row in rows]


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv_rows(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def sensor_key(row: dict[str, Any]) -> tuple[str, str]:
    return (str(row.get("sn", "")).strip(), str(row.get("device_name", "")).strip())


def read_point_config(path: Path) -> dict[tuple[str, str], list[dict[str, Any]]]:
    if not path.exists():
        return {}
    config_rows = read_csv_rows(path)
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for raw in config_rows:
        if str(raw.get("enabled", "1")).strip().lower() in {"0", "false", "no", "否"}:
            continue
        x = to_float(raw.get("target_x_px"))
        y = to_float(raw.get("target_y_px"))
        if x is None or y is None:
            continue
        point_order = to_float(raw.get("point_order"))
        radius = to_float(raw.get("search_radius_px"))
        row = dict(raw)
        row["_point_order"] = int(point_order or (len(grouped[sensor_key(row)]) + 1))
        row["_target_x"] = float(x)
        row["_target_y"] = float(y)
        row["_search_radius"] = float(radius or 220.0)
        grouped[sensor_key(row)].append(row)
    for rows in grouped.values():
        rows.sort(key=lambda item: int(item["_point_order"]))
    return grouped


def write_point_config(path: Path, rows: list[dict[str, Any]]) -> None:
    normalized: list[dict[str, Any]] = []
    for row in rows:
        if to_float(row.get("target_x_px")) is None or to_float(row.get("target_y_px")) is None:
            continue
        point_order = int(to_float(row.get("point_order")) or 0)
        normalized.append(
            {
                "sn": row.get("sn", ""),
                "device_name": row.get("device_name", ""),
                "point_order": point_order,
                "point_name": row.get("point_name") or f"P{point_order}",
                "target_x_px": round_if_number(row.get("target_x_px"), 3),
                "target_y_px": round_if_number(row.get("target_y_px"), 3),
                "search_radius_px": round_if_number(row.get("search_radius_px") or 220, 3),
                "enabled": row.get("enabled", "1"),
                "note": row.get("note", ""),
                "updated_from_image": row.get("updated_from_image", ""),
            }
        )
    normalized.sort(
        key=lambda row: (
            str(row.get("sn", "")),
            str(row.get("device_name", "")),
            int(to_float(row.get("point_order")) or 0),
        )
    )
    write_csv_rows(path, normalized, POINT_CONFIG_FIELDS)


def infer_point_config_from_points(rows: list[dict[str, Any]], search_radius_px: float = 220.0) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in normalize_point_rows(rows):
        x = to_float(row.get("x_px"))
        y = to_float(row.get("y_px"))
        point_order = to_float(row.get("point_order"))
        if x is None or y is None or point_order is None:
            continue
        key = (str(row.get("sn", "")), str(row.get("device_name", "")), int(point_order))
        grouped[key].append(row)
    config_rows: list[dict[str, Any]] = []
    for (sn, device_name, point_order), group in sorted(grouped.items()):
        xs = [float(row["x_px"]) for row in group if to_float(row.get("x_px")) is not None]
        ys = [float(row["y_px"]) for row in group if to_float(row.get("y_px")) is not None]
        if not xs or not ys:
            continue
        first = group[0]
        config_rows.append(
            {
                "sn": sn,
                "device_name": device_name,
                "point_order": point_order,
                "point_name": f"P{point_order}",
                "target_x_px": statistics.median(xs),
                "target_y_px": statistics.median(ys),
                "search_radius_px": search_radius_px,
                "enabled": "1",
                "note": "由既有测点中位数初始化，建议人工复核后保存",
                "updated_from_image": first.get("filename", ""),
            }
        )
    return config_rows


def source_profile_csv(row: dict[str, Any]) -> Path:
    return Path(str(row["output_dir"])) / f"{Path(str(row['filename'])).stem}_profile.csv"


def review_fractions(points_per_image: int) -> list[float]:
    if points_per_image <= 0:
        raise ValueError("points_per_image must be positive")
    if points_per_image == 3:
        return [0.2, 0.5, 0.8]
    if points_per_image == 5:
        return [0.1, 0.3, 0.5, 0.7, 0.9]
    step = 1.0 / (points_per_image + 1)
    return [step * (idx + 1) for idx in range(points_per_image)]


def valid_profile_rows(profile_rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for raw in profile_rows:
        distance = to_float(raw.get("distance_px"))
        width_mm = to_float(raw.get("profile_width_mm"))
        x = to_float(raw.get("x"))
        y = to_float(raw.get("y"))
        if distance is None or width_mm is None or x is None or y is None:
            continue
        row = dict(raw)
        row["_distance_px"] = distance
        row["_profile_width_mm"] = width_mm
        row["_x"] = x
        row["_y"] = y
        rows.append(row)
    rows.sort(key=lambda item: item["_distance_px"])
    return rows


def select_review_points(profile_rows: list[dict[str, str]], points_per_image: int) -> list[tuple[float, dict[str, Any]]]:
    rows = valid_profile_rows(profile_rows)
    if not rows:
        return []
    fractions = review_fractions(points_per_image)
    if len(rows) <= points_per_image:
        if len(rows) == 1:
            return [(0.5, rows[0])]
        return [(idx / (len(rows) - 1), row) for idx, row in enumerate(rows)]

    min_distance = rows[0]["_distance_px"]
    max_distance = rows[-1]["_distance_px"]
    used: set[int] = set()
    selected: list[tuple[float, dict[str, Any]]] = []
    for fraction in fractions:
        if max_distance > min_distance:
            target = min_distance + fraction * (max_distance - min_distance)
            candidates = sorted(
                enumerate(rows),
                key=lambda item: (abs(item[1]["_distance_px"] - target), item[0]),
            )
        else:
            target_index = round(fraction * (len(rows) - 1))
            candidates = sorted(enumerate(rows), key=lambda item: (abs(item[0] - target_index), item[0]))
        for idx, row in candidates:
            if idx not in used:
                used.add(idx)
                selected.append((fraction, row))
                break
    selected.sort(key=lambda item: item[1]["_distance_px"])
    return selected


def read_image_bgr(path: Path):
    import cv2
    import numpy as np

    data = np.fromfile(str(path), dtype=np.uint8)
    return cv2.imdecode(data, cv2.IMREAD_COLOR)


def platform_marker_mask(image_bgr, *, dilate_px: int = 24):
    """Detect residual platform labels/guide lines, not concrete cracks."""
    import cv2
    import numpy as np

    height, width = image_bgr.shape[:2]
    gray = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2GRAY)
    hsv = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2HSV)
    hue, sat, val = cv2.split(hsv)

    # Saturated blue/purple/red guide lines are not natural concrete texture.
    colored = (
        (sat > 70)
        & (val > 80)
        & (
            ((hue >= 95) & (hue <= 160))
            | (hue <= 8)
            | (hue >= 170)
        )
    ).astype(np.uint8) * 255

    dark = (gray < 70).astype(np.uint8) * 255
    bright = ((gray > 165) & (sat < 85)).astype(np.uint8) * 255

    # Group white digits/text, then retain groups sitting on a dark label.
    text_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (13, 5))
    bright_groups = cv2.dilate(bright, text_kernel, iterations=1)
    num, labels, stats, _centers = cv2.connectedComponentsWithStats(bright_groups, 8)
    label_mask = np.zeros_like(gray, dtype=np.uint8)
    for idx in range(1, num):
        x = int(stats[idx, cv2.CC_STAT_LEFT])
        y = int(stats[idx, cv2.CC_STAT_TOP])
        w = int(stats[idx, cv2.CC_STAT_WIDTH])
        h = int(stats[idx, cv2.CC_STAT_HEIGHT])
        area = int(stats[idx, cv2.CC_STAT_AREA])
        if area < 15 or w < 8 or h < 5 or w > 260 or h > 80:
            continue
        x0 = max(0, x - 12)
        y0 = max(0, y - 12)
        x1 = min(width, x + w + 12)
        y1 = min(height, y + h + 12)
        region = dark[y0:y1, x0:x1]
        if region.size == 0:
            continue
        dark_fraction = float(np.count_nonzero(region)) / float(region.size)
        # White text on black blocks has a high local dark fraction. Normal
        # concrete speckles rarely satisfy this after grouping.
        if dark_fraction >= 0.08:
            label_mask[y0:y1, x0:x1] = 255

    # Timestamps are generally white text on a black strip in the top-left.
    top_h = max(28, int(height * 0.045))
    left_w = max(220, int(width * 0.22))
    top_left = gray[:top_h, :left_w]
    if top_left.size and float(np.mean(top_left < 90)) > 0.12 and np.count_nonzero(top_left > 160) > 20:
        label_mask[: top_h + 8, : left_w + 8] = 255

    mask = cv2.bitwise_or(colored, label_mask)
    kernel_size = max(5, int(dilate_px) | 1)
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (kernel_size, kernel_size))
    mask = cv2.dilate(mask, kernel, iterations=1)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=1)
    return mask


def row_xy(row: dict[str, Any]) -> tuple[float | None, float | None]:
    return to_float(row.get("x")), to_float(row.get("y"))


def line_contamination_score(mask, row: dict[str, Any], *, radius_px: int = 38) -> float:
    import cv2
    import numpy as np

    height, width = mask.shape[:2]
    x = to_float(row.get("x"))
    y = to_float(row.get("y"))
    if x is None or y is None:
        return 1.0
    cx = int(round(x))
    cy = int(round(y))
    x0 = max(0, cx - radius_px)
    y0 = max(0, cy - radius_px)
    x1 = min(width, cx + radius_px + 1)
    y1 = min(height, cy + radius_px + 1)
    patch = mask[y0:y1, x0:x1]
    patch_hit = 1.0 if np.count_nonzero(patch) > 0 else 0.0

    lx = to_float(row.get("left_x"))
    ly = to_float(row.get("left_y"))
    rx = to_float(row.get("right_x"))
    ry = to_float(row.get("right_y"))
    if None in (lx, ly, rx, ry):
        return patch_hit
    line_mask = np.zeros_like(mask, dtype=np.uint8)
    cv2.line(
        line_mask,
        (int(round(lx)), int(round(ly))),
        (int(round(rx)), int(round(ry))),
        255,
        max(3, radius_px // 4),
    )
    line_hit = 1.0 if np.count_nonzero((line_mask > 0) & (mask > 0)) > 0 else 0.0
    return max(patch_hit, line_hit)


def local_marker_contamination_score(image_bgr, row: dict[str, Any], *, radius_px: int = 58) -> float:
    """Fast local check around one candidate point/normal line."""
    import cv2
    import numpy as np

    if image_bgr is None:
        return 0.0
    height, width = image_bgr.shape[:2]
    coords = [to_float(row.get(key)) for key in ("x", "y", "left_x", "left_y", "right_x", "right_y")]
    if coords[0] is None or coords[1] is None:
        return 1.0
    valid = [(float(coords[i]), float(coords[i + 1])) for i in range(0, len(coords), 2) if coords[i] is not None and coords[i + 1] is not None]
    xs = [xy[0] for xy in valid]
    ys = [xy[1] for xy in valid]
    x0 = max(0, int(round(min(xs) - radius_px)))
    y0 = max(0, int(round(min(ys) - radius_px)))
    x1 = min(width, int(round(max(xs) + radius_px + 1)))
    y1 = min(height, int(round(max(ys) + radius_px + 1)))
    if x1 <= x0 or y1 <= y0:
        return 1.0
    patch = image_bgr[y0:y1, x0:x1]
    gray = cv2.cvtColor(patch, cv2.COLOR_BGR2GRAY)
    hsv = cv2.cvtColor(patch, cv2.COLOR_BGR2HSV)
    hue, sat, val = cv2.split(hsv)

    colored = (
        (sat > 70)
        & (val > 80)
        & (
            ((hue >= 95) & (hue <= 160))
            | (hue <= 8)
            | (hue >= 170)
        )
    )
    colored_count = int(np.count_nonzero(colored))
    if colored_count >= 8:
        return 1.0

    dark = gray < 70
    bright = (gray > 165) & (sat < 90)
    dark_count = int(np.count_nonzero(dark))
    bright_count = int(np.count_nonzero(bright))
    area = int(gray.size)
    if area <= 0:
        return 1.0

    # Residual width labels are black rectangles with white digits. Natural
    # concrete texture may be dark or bright, but rarely contains enough white
    # text pixels embedded in a compact black block around the crack.
    dark_fraction = dark_count / area
    if bright_count >= 45 and dark_fraction >= 0.12:
        bright_u8 = bright.astype(np.uint8) * 255
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (9, 5))
        grouped = cv2.dilate(bright_u8, kernel, iterations=1)
        num, _labels, stats, _centers = cv2.connectedComponentsWithStats(grouped, 8)
        for idx in range(1, num):
            bw = int(stats[idx, cv2.CC_STAT_WIDTH])
            bh = int(stats[idx, cv2.CC_STAT_HEIGHT])
            ba = int(stats[idx, cv2.CC_STAT_AREA])
            if ba >= 30 and 10 <= bw <= 180 and 5 <= bh <= 55:
                return 1.0
    return 0.0


def select_clean_review_points(
    profile_rows: list[dict[str, str]],
    points_per_image: int,
    marker_mask=None,
    image_bgr=None,
) -> tuple[list[dict[str, Any]], str]:
    rows = valid_profile_rows(profile_rows)
    if not rows:
        return [], "no_valid_profiles"
    fractions = review_fractions(points_per_image)
    min_distance = rows[0]["_distance_px"]
    max_distance = rows[-1]["_distance_px"]
    path_span = max(1.0, float(max_distance - min_distance))
    max_shift_px = max(40.0, path_span * 0.18)
    min_separation_px = max(35.0, path_span / float(points_per_image + 1) * 0.45)
    used: set[int] = set()
    selected_distances: list[float] = []
    selected: list[dict[str, Any]] = []
    for fraction in fractions:
        if max_distance > min_distance:
            target = min_distance + fraction * (max_distance - min_distance)
        else:
            target = rows[round(fraction * (len(rows) - 1))]["_distance_px"]
        candidates = sorted(
            enumerate(rows),
            key=lambda item: (abs(item[1]["_distance_px"] - target), item[0]),
        )
        original_idx, original_row = candidates[0]
        original_dirty = candidate_is_contaminated(original_row, marker_mask, image_bgr)
        chosen_idx: int | None = None
        chosen_row: dict[str, Any] | None = None
        for idx, row in candidates:
            if idx in used:
                continue
            shift = abs(float(row["_distance_px"]) - float(target))
            if shift > max_shift_px:
                continue
            if any(abs(float(row["_distance_px"]) - prev) < min_separation_px for prev in selected_distances):
                continue
            dirty = candidate_is_contaminated(row, marker_mask, image_bgr)
            if not dirty:
                chosen_idx = idx
                chosen_row = row
                break
        if chosen_row is None:
            continue
        used.add(int(chosen_idx))
        selected_distances.append(float(chosen_row["_distance_px"]))
        replacement_px = abs(float(chosen_row["_distance_px"]) - float(original_row["_distance_px"]))
        selected.append(
            {
                "target_fraction": fraction,
                "row": chosen_row,
                "target_distance_px": target,
                "original_contaminated": bool(original_dirty),
                "replacement_from_target_px": replacement_px,
                "selection_reason": (
                    "target_replaced_due_to_marker"
                    if original_dirty and chosen_idx != original_idx
                    else "clean_target"
                ),
            }
        )
    selected.sort(key=lambda item: item["row"]["_distance_px"])
    if len(selected) < points_per_image:
        return selected, "insufficient_clean_points"
    return selected, ""


def select_configured_review_points(
    profile_rows: list[dict[str, str]],
    config_rows: list[dict[str, Any]],
    marker_mask=None,
    image_bgr=None,
) -> tuple[list[dict[str, Any]], str]:
    rows = valid_profile_rows(profile_rows)
    if not rows:
        return [
            {
                "point_order": int(config["_point_order"]),
                "point_name": config.get("point_name") or f"P{int(config['_point_order'])}",
                "target_fraction": "",
                "target_x_px": float(config["_target_x"]),
                "target_y_px": float(config["_target_y"]),
                "target_search_radius_px": float(config["_search_radius"]),
                "row": None,
                "target_distance_px": "",
                "original_contaminated": False,
                "replacement_from_target_px": "",
                "selection_reason": "sensor_config_no_valid_profile",
                "point_config_source": "sensor_point_config",
                "exclude_reason": "no_valid_profiles",
            }
            for config in config_rows
        ], ""
    used: set[int] = set()
    selected: list[dict[str, Any]] = []
    for config in config_rows:
        target_x = float(config["_target_x"])
        target_y = float(config["_target_y"])
        radius = float(config["_search_radius"])
        candidates = sorted(
            enumerate(rows),
            key=lambda item: (
                (float(item[1]["_x"]) - target_x) ** 2 + (float(item[1]["_y"]) - target_y) ** 2,
                item[0],
            ),
        )
        chosen_idx: int | None = None
        chosen_row: dict[str, Any] | None = None
        nearest_distance = ""
        for idx, row in candidates:
            if idx in used:
                continue
            distance = math.hypot(float(row["_x"]) - target_x, float(row["_y"]) - target_y)
            if nearest_distance == "":
                nearest_distance = distance
            if distance > radius:
                continue
            if candidate_is_contaminated(row, marker_mask, image_bgr):
                continue
            chosen_idx = idx
            chosen_row = row
            nearest_distance = distance
            break
        if chosen_row is None:
            selected.append(
                {
                    "point_order": int(config["_point_order"]),
                    "point_name": config.get("point_name") or f"P{int(config['_point_order'])}",
                    "target_fraction": "",
                    "target_x_px": target_x,
                    "target_y_px": target_y,
                    "target_search_radius_px": radius,
                    "row": None,
                    "target_distance_px": "",
                    "original_contaminated": False,
                    "replacement_from_target_px": nearest_distance,
                    "selection_reason": "sensor_config_no_clean_profile",
                    "point_config_source": "sensor_point_config",
                    "exclude_reason": f"no_clean_profile_near_config_P{int(config['_point_order'])}",
                }
            )
            continue
        used.add(int(chosen_idx))
        selected.append(
            {
                "point_order": int(config["_point_order"]),
                "point_name": config.get("point_name") or f"P{int(config['_point_order'])}",
                "target_fraction": "",
                "target_x_px": target_x,
                "target_y_px": target_y,
                "target_search_radius_px": radius,
                "row": chosen_row,
                "target_distance_px": "",
                "original_contaminated": False,
                "replacement_from_target_px": nearest_distance,
                "selection_reason": "sensor_config_nearest_profile",
                "point_config_source": "sensor_point_config",
                "exclude_reason": "",
            }
        )
    selected.sort(key=lambda item: int(item["point_order"]))
    return selected, ""


def candidate_is_contaminated(row: dict[str, Any], marker_mask=None, image_bgr=None) -> bool:
    if image_bgr is not None:
        return local_marker_contamination_score(image_bgr, row) > 0
    if marker_mask is not None:
        return line_contamination_score(marker_mask, row) > 0
    return False


def actual_position_pct(profile_row: dict[str, Any], all_profile_rows: list[dict[str, str]]) -> float | str:
    rows = valid_profile_rows(all_profile_rows)
    if len(rows) < 2:
        return ""
    min_distance = rows[0]["_distance_px"]
    max_distance = rows[-1]["_distance_px"]
    if max_distance <= min_distance:
        return ""
    return 100.0 * (profile_row["_distance_px"] - min_distance) / (max_distance - min_distance)


def build_point_rows(
    summary_csv: Path,
    points_per_image: int,
    *,
    avoid_markers: bool = True,
    excluded_csv: Path | None = None,
    point_config_csv: Path | None = None,
) -> list[dict[str, Any]]:
    summary_rows = read_csv_rows(summary_csv)
    config_map = read_point_config(point_config_csv) if point_config_csv and point_config_csv.exists() else {}
    point_rows: list[dict[str, Any]] = []
    excluded_rows: list[dict[str, Any]] = []
    for image_idx, image_row in enumerate(summary_rows, start=1):
        profile_csv = source_profile_csv(image_row)
        if not profile_csv.exists():
            excluded_rows.append(
                {
                    "image_record_id": image_idx,
                    "sn": image_row.get("sn", ""),
                    "device_name": image_row.get("device_name", ""),
                    "uptime": image_row.get("uptime", ""),
                    "filename": image_row.get("filename", ""),
                    "input_image_path": image_row.get("input_path", ""),
                    "exclude_reason": "missing_profile_csv",
                }
            )
            continue
        profile_rows = read_csv_rows(profile_csv)
        marker_mask = None
        image_bgr = None
        if avoid_markers:
            image_bgr = read_image_bgr(Path(str(image_row.get("input_path", ""))))
        config_rows = config_map.get(sensor_key(image_row), [])
        if config_rows:
            selected, exclude_reason = select_configured_review_points(
                profile_rows,
                config_rows,
                marker_mask=marker_mask,
                image_bgr=image_bgr,
            )
            required_points = len(config_rows)
            selection_mode = "sensor_config"
        else:
            selected, exclude_reason = select_clean_review_points(
                profile_rows,
                points_per_image,
                marker_mask=marker_mask,
                image_bgr=image_bgr,
            )
            required_points = points_per_image
            selection_mode = "auto_fraction"
        if exclude_reason:
            excluded_rows.append(
                {
                    "image_record_id": image_idx,
                    "sn": image_row.get("sn", ""),
                    "device_name": image_row.get("device_name", ""),
                    "uptime": image_row.get("uptime", ""),
                    "filename": image_row.get("filename", ""),
                    "input_image_path": image_row.get("input_path", ""),
                    "selected_clean_points": len(selected),
                    "required_points": required_points,
                    "exclude_reason": exclude_reason,
                }
            )
            continue
        for fallback_order, selection in enumerate(selected, start=1):
            point_order = int(selection.get("point_order") or fallback_order)
            target_fraction = to_float(selection.get("target_fraction"))
            point = selection.get("row")
            point_invalid_reason = str(selection.get("exclude_reason", "") or "")
            dirty = bool(point_invalid_reason) or (
                candidate_is_contaminated(point, marker_mask, image_bgr) if point is not None else True
            )
            point_rows.append(
                {
                    "point_id": f"{image_idx:04d}-P{point_order}",
                    "image_record_id": image_idx,
                    "point_order": point_order,
                    "target_position_pct": round(target_fraction * 100.0, 3) if target_fraction is not None else "",
                    "actual_position_pct": round_if_number(actual_position_pct(point, profile_rows), 3) if point is not None else "",
                    "sn": image_row.get("sn", ""),
                    "device_name": image_row.get("device_name", ""),
                    "uptime": image_row.get("uptime", ""),
                    "filename": image_row.get("filename", ""),
                    "input_image_path": image_row.get("input_path", ""),
                    "overlay_path": image_row.get("overlay_path", ""),
                    "profile_csv_path": str(profile_csv),
                    "target_x_px": round_if_number(selection.get("target_x_px"), 3),
                    "target_y_px": round_if_number(selection.get("target_y_px"), 3),
                    "target_search_radius_px": round_if_number(selection.get("target_search_radius_px"), 3),
                    "x_px": round_if_number(point.get("_x"), 3) if point is not None else "",
                    "y_px": round_if_number(point.get("_y"), 3) if point is not None else "",
                    "left_x_px": round_if_number(to_float(point.get("left_x")), 3) if point is not None else "",
                    "left_y_px": round_if_number(to_float(point.get("left_y")), 3) if point is not None else "",
                    "right_x_px": round_if_number(to_float(point.get("right_x")), 3) if point is not None else "",
                    "right_y_px": round_if_number(to_float(point.get("right_y")), 3) if point is not None else "",
                    "auto_width_mm": round_if_number(point.get("_profile_width_mm"), 6) if point is not None else "",
                    "auto_width_px": round_if_number(to_float(point.get("profile_width_px")), 6) if point is not None else "",
                    "mask_width_mm": round_if_number(to_float(point.get("mask_width_mm")), 6) if point is not None else "",
                    "contrast": round_if_number(to_float(point.get("contrast")), 6) if point is not None else "",
                    "threshold": round_if_number(to_float(point.get("threshold")), 6) if point is not None else "",
                    "quality_usable": image_row.get("quality_usable", ""),
                    "marker_contaminated": int(bool(dirty)),
                    "selection_reason": selection.get("selection_reason", ""),
                    "replacement_from_target_px": round_if_number(
                        selection.get("replacement_from_target_px"), 3
                    ),
                    "point_config_source": selection.get("point_config_source", selection_mode),
                    "review_usable": int(not dirty),
                    "exclude_reason": point_invalid_reason,
                    "image_deleted": "0",
                    "delete_reason": "",
                    "manual_width_mm": format_manual_width(point.get("_profile_width_mm")) if point is not None else "",
                    "manual_source": MANUAL_PREFILL_SOURCE if point is not None else "",
                    "review_status": "pending" if point is not None else "skipped",
                    "manual_note": point_invalid_reason,
                }
            )
    if excluded_csv is not None:
        excluded_fields = [
            "image_record_id",
            "sn",
            "device_name",
            "uptime",
            "filename",
            "input_image_path",
            "selected_clean_points",
            "required_points",
            "exclude_reason",
        ]
        write_csv_rows(excluded_csv, excluded_rows, excluded_fields)
    return point_rows


def round_if_number(value: Any, digits: int) -> Any:
    number = to_float(value)
    if number is None:
        return value if value is not None else ""
    return round(number, digits)


def create_point_excel(csv_path: Path, xlsx_path: Path) -> Path:
    rows = normalize_point_rows(read_csv_rows(csv_path))
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "测点复核录入"
    params = wb.create_sheet("参数")
    info = wb.create_sheet("说明")

    params_rows = [
        ["参数", "值", "说明"],
        ["abs_tol_mm", 0.02, "绝对误差阈值，默认 0.02 mm，可按期刊或工程要求调整"],
        ["rel_tol", 0.15, "相对误差阈值，默认 15%"],
        ["qualified_rule", "abs_error <= MAX(abs_tol_mm, rel_tol * manual_width_mm)", "符合度判定规则；仅统计 review_status=reviewed 且未删除的测点"],
        ["manual_instrument", f"{MANUAL_INSTRUMENT_NAME} {MANUAL_INSTRUMENT_MODEL}", "人工测读仪器"],
        ["manufacturer", MANUAL_INSTRUMENT_MANUFACTURER, "制造厂商"],
        ["range", MANUAL_INSTRUMENT_RANGE, "测量范围"],
        ["uncertainty", MANUAL_INSTRUMENT_UNCERTAINTY, "校准证书给出的示值误差测量不确定度"],
        ["basis", MANUAL_INSTRUMENT_BASIS, "校准依据"],
    ]
    for row in params_rows:
        params.append(row)
    for cell in params[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center")
    params.freeze_panes = "A2"
    params.column_dimensions["A"].width = 18
    params.column_dimensions["B"].width = 20
    params.column_dimensions["C"].width = 78

    info_rows = [
        ["字段", "含义"],
        ["point_id", "测点编号，一张图内按主裂缝弧长位置依次编号"],
        ["target_position_pct", "自动等距取点时的目标弧长位置；使用传感器配置时可为空"],
        ["actual_position_pct", "程序在已有有效法向剖面中就近匹配后的实际弧长位置"],
        ["target_x_px", "传感器级人工配置测点的目标 x 坐标，单位 px"],
        ["target_y_px", "传感器级人工配置测点的目标 y 坐标，单位 px"],
        ["target_search_radius_px", "围绕人工配置测点搜索有效裂缝剖面的半径，单位 px"],
        ["auto_width_mm", "该测点程序自动识别宽度"],
        ["point_config_source", "sensor_point_config 表示按传感器人工配置取点；auto_fraction 表示自动等距取点"],
        ["actual_override_source", "manual_actual_override 表示该图该测点的实际测宽剖面曾由人工点击调整"],
        ["actual_override_distance_px", "人工点击位置与吸附到的有效法向剖面中心距离，单位 px"],
        ["original_*", "人工调整实际测点前保留的原自动测点坐标、宽度和选点原因，用于追溯或恢复"],
        ["image_deleted", "图片级软删除标记；1 表示该图不参与指标统计，原图文件不物理删除"],
        ["delete_reason", "图片级删除原因"],
        ["manual_width_mm", "人工在同一编号测点读取的缝宽，单位 mm；默认按程序值预填并保留 2 位小数"],
        ["manual_source", "manual_confirmed 表示逐点人工确认；prefilled_auto_2dp 表示仅为默认预填；batch_confirmed_* 表示批量确认"],
        ["review_status", "pending/reviewed/skipped/deleted；只有 reviewed 且 image_deleted=0 的测点参与指标统计"],
        ["manual_note", "人工备注，例如图像模糊、测点不可读、裂缝边界不清等"],
    ]
    for row in info_rows:
        info.append(row)
    for cell in info[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center")
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 88

    fields = POINT_FIELDS + ["abs_error_mm", "qualified"]
    ws.append(fields)
    manual_col = fields.index("manual_width_mm") + 1
    auto_col = fields.index("auto_width_mm") + 1
    status_col = fields.index("review_status") + 1
    deleted_col = fields.index("image_deleted") + 1
    abs_col = fields.index("abs_error_mm") + 1
    qualified_col = fields.index("qualified") + 1

    for row_idx, row in enumerate(rows, start=2):
        values = [row.get(field, "") for field in POINT_FIELDS]
        manual_cell = f"{get_column_letter(manual_col)}{row_idx}"
        auto_cell = f"{get_column_letter(auto_col)}{row_idx}"
        status_cell = f"{get_column_letter(status_col)}{row_idx}"
        deleted_cell = f"{get_column_letter(deleted_col)}{row_idx}"
        abs_cell = f"{get_column_letter(abs_col)}{row_idx}"
        values.append(f'=IF(OR({manual_cell}="",{status_cell}<>"reviewed",{deleted_cell}=1),"",ABS({auto_cell}-{manual_cell}))')
        values.append(f'=IF({abs_cell}="","",--({abs_cell}<=MAX(参数!$B$2,参数!$B$3*{manual_cell})))')
        ws.append(values)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    table = Table(displayName="ManualPointReviewTable", ref=f"A1:{get_column_letter(len(fields))}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)

    widths = {
        "A": 13,
        "B": 15,
        "C": 10,
        "D": 18,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 20,
        "I": 28,
        "J": 45,
        "K": 45,
        "L": 45,
        "M": 10,
        "N": 10,
        "O": 12,
        "P": 12,
        "Q": 12,
        "R": 12,
        "S": 16,
        "T": 14,
        "U": 14,
        "V": 12,
        "W": 12,
        "X": 13,
        "Y": 16,
        "Z": 18,
        "AA": 42,
        "AB": 14,
        "AC": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    input_col = fields.index("input_image_path") + 1
    overlay_col = fields.index("overlay_path") + 1
    profile_col = fields.index("profile_csv_path") + 1
    source_col = fields.index("manual_source") + 1
    deleted_col = fields.index("image_deleted") + 1
    delete_reason_col = fields.index("delete_reason") + 1
    note_col = fields.index("manual_note") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.protection = Protection(locked=True)
        for col_idx in [input_col, overlay_col, profile_col]:
            cell = row[col_idx - 1]
            cell.hyperlink = cell.value
            cell.font = Font(color="0563C1", underline="single")
        for col_idx, field in enumerate(fields, start=1):
            if field.endswith("_px") or field.endswith("_mm") or field.endswith("_pct") or field in {
                "contrast",
                "threshold",
                "replacement_from_target_px",
                "abs_error_mm",
            }:
                row[col_idx - 1].number_format = "0.0000"
        row[manual_col - 1].number_format = "0.00"
        for col_idx in [manual_col, source_col, status_col, deleted_col, delete_reason_col, note_col]:
            row[col_idx - 1].protection = Protection(locked=False)
        row[qualified_col - 1].number_format = "0"

    wb.save(xlsx_path)
    return xlsx_path


def read_point_rows(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".xlsx":
        wb = load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [str(cell.value).strip() for cell in ws[1]]
        rows = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            if row.get("point_id"):
                rows.append(row)
        return normalize_point_rows(rows)
    return normalize_point_rows(read_csv_rows(path))


def is_deleted(row: dict[str, Any]) -> bool:
    status = str(row.get("review_status", "")).strip().lower()
    return is_truthy(row.get("image_deleted")) or status in {"deleted", "image_deleted", "删除", "已删除"}


def is_skipped(row: dict[str, Any]) -> bool:
    status = str(row.get("review_status", "")).strip().lower()
    usable = str(row.get("review_usable", "1")).strip().lower()
    return is_deleted(row) or status in {"skip", "skipped", "跳过", "剔除", "invalid"} or usable in {"0", "false", "no"}


def is_reviewed(row: dict[str, Any]) -> bool:
    status = str(row.get("review_status", "")).strip().lower()
    return status in {"reviewed", "confirmed", "manual_confirmed"} and not is_skipped(row)


def width_bin(width: float | None) -> str:
    if width is None:
        return "未填写"
    if width < 0.05:
        return "<0.05 mm"
    if width < 0.10:
        return "0.05-0.10 mm"
    return ">=0.10 mm"


def metric_summary(error_rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not error_rows:
        return {
            "n": 0,
            "mae_mm": "",
            "rmse_mm": "",
            "bias_mm": "",
            "median_abs_error_mm": "",
            "max_abs_error_mm": "",
            "qualified_rate_pct": "",
        }
    errors = [float(row["error_mm"]) for row in error_rows]
    abs_errors = [float(row["abs_error_mm"]) for row in error_rows]
    return {
        "n": len(error_rows),
        "mae_mm": sum(abs_errors) / len(abs_errors),
        "rmse_mm": math.sqrt(sum(error * error for error in errors) / len(errors)),
        "bias_mm": sum(errors) / len(errors),
        "median_abs_error_mm": statistics.median(abs_errors),
        "max_abs_error_mm": max(abs_errors),
        "qualified_rate_pct": 100.0 * sum(int(row["qualified"]) for row in error_rows) / len(error_rows),
    }


def rounded(value: Any) -> Any:
    return round(value, 6) if isinstance(value, float) else value


def evaluate_points_file(input_path: Path, out_dir: Path, abs_tol_mm: float, rel_tol: float) -> dict[str, Path]:
    rows = read_point_rows(input_path)
    out_dir.mkdir(parents=True, exist_ok=True)

    error_rows: list[dict[str, Any]] = []
    for row in rows:
        if not is_reviewed(row):
            continue
        manual = to_float(row.get("manual_width_mm"))
        auto = to_float(row.get("auto_width_mm"))
        if manual is None or auto is None:
            continue
        error = auto - manual
        abs_error = abs(error)
        threshold = max(abs_tol_mm, rel_tol * manual)
        record = {
            "point_id": row.get("point_id", ""),
            "image_record_id": row.get("image_record_id", ""),
            "point_order": row.get("point_order", ""),
            "target_position_pct": row.get("target_position_pct", ""),
            "actual_position_pct": row.get("actual_position_pct", ""),
            "actual_override_source": row.get("actual_override_source", ""),
            "sn": row.get("sn", ""),
            "device_name": row.get("device_name", ""),
            "uptime": row.get("uptime", ""),
            "filename": row.get("filename", ""),
            "auto_width_mm": auto,
            "manual_width_mm": manual,
            "error_mm": error,
            "abs_error_mm": abs_error,
            "threshold_mm": threshold,
            "qualified": int(abs_error <= threshold),
            "width_bin": width_bin(manual),
            "review_status": row.get("review_status", ""),
            "manual_source": row.get("manual_source", ""),
            "image_deleted": row.get("image_deleted", ""),
            "delete_reason": row.get("delete_reason", ""),
            "manual_note": row.get("manual_note", ""),
            "input_image_path": row.get("input_image_path", ""),
        }
        error_rows.append(record)

    point_errors_csv = out_dir / "manual_point_errors.csv"
    point_error_fields = [
        "point_id",
        "image_record_id",
        "point_order",
        "target_position_pct",
        "actual_position_pct",
        "actual_override_source",
        "sn",
        "device_name",
        "uptime",
        "filename",
        "auto_width_mm",
        "manual_width_mm",
        "error_mm",
        "abs_error_mm",
        "threshold_mm",
        "qualified",
        "width_bin",
        "review_status",
        "manual_source",
        "image_deleted",
        "delete_reason",
        "manual_note",
        "input_image_path",
    ]
    write_csv_rows(point_errors_csv, error_rows, point_error_fields)

    image_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in error_rows:
        image_groups[str(row["image_record_id"])].append(row)
    image_rows: list[dict[str, Any]] = []
    for image_id, group in sorted(image_groups.items(), key=lambda item: int(item[0]) if str(item[0]).isdigit() else str(item[0])):
        auto_values = [float(row["auto_width_mm"]) for row in group]
        manual_values = [float(row["manual_width_mm"]) for row in group]
        auto_mean = statistics.fmean(auto_values)
        manual_mean = statistics.fmean(manual_values)
        error = auto_mean - manual_mean
        threshold = max(abs_tol_mm, rel_tol * manual_mean)
        first = group[0]
        image_rows.append(
            {
                "image_record_id": image_id,
                "sn": first["sn"],
                "device_name": first["device_name"],
                "uptime": first["uptime"],
                "filename": first["filename"],
                "reviewed_points": len(group),
                "auto_mean_width_mm": auto_mean,
                "manual_mean_width_mm": manual_mean,
                "error_mm": error,
                "abs_error_mm": abs(error),
                "threshold_mm": threshold,
                "qualified": int(abs(error) <= threshold),
                "input_image_path": first["input_image_path"],
            }
        )

    image_errors_csv = out_dir / "manual_image_errors.csv"
    image_fields = [
        "image_record_id",
        "sn",
        "device_name",
        "uptime",
        "filename",
        "reviewed_points",
        "auto_mean_width_mm",
        "manual_mean_width_mm",
        "error_mm",
        "abs_error_mm",
        "threshold_mm",
        "qualified",
        "input_image_path",
    ]
    write_csv_rows(image_errors_csv, image_rows, image_fields)

    metric_rows: list[dict[str, Any]] = []
    groups: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in error_rows:
        groups[("测点级-总体", "全部")].append(row)
        groups[("测点级-构件", str(row["device_name"]))].append(row)
        groups[("测点级-位置", f"P{row['point_order']}")].append(row)
        groups[("测点级-宽度区间", row["width_bin"])].append(row)
    for row in image_rows:
        groups[("图像级均值-总体", "全部")].append(row)
        groups[("图像级均值-构件", str(row["device_name"]))].append(row)
    for (group_type, group_name), group_rows in sorted(groups.items()):
        summary = metric_summary(group_rows)
        metric_rows.append(
            {
                "group_type": group_type,
                "group_name": group_name,
                **{key: rounded(value) for key, value in summary.items()},
            }
        )

    metrics_csv = out_dir / "manual_point_metrics.csv"
    metric_fields = [
        "group_type",
        "group_name",
        "n",
        "mae_mm",
        "rmse_mm",
        "bias_mm",
        "median_abs_error_mm",
        "max_abs_error_mm",
        "qualified_rate_pct",
    ]
    write_csv_rows(metrics_csv, metric_rows, metric_fields)

    paper_rows = [
        {
            "评价对象": row["group_type"],
            "分组": row["group_name"],
            "样本数": row["n"],
            "MAE/mm": row["mae_mm"],
            "RMSE/mm": row["rmse_mm"],
            "中位绝对误差/mm": row["median_abs_error_mm"],
            "最大绝对误差/mm": row["max_abs_error_mm"],
            "符合度/%": row["qualified_rate_pct"],
            "备注": "由测点级人工复核值自动计算",
        }
        for row in metric_rows
        if row["group_type"] in {"测点级-总体", "图像级均值-总体"}
    ]
    paper_csv = out_dir / "paper_metric_table.csv"
    write_csv_rows(
        paper_csv,
        paper_rows,
        ["评价对象", "分组", "样本数", "MAE/mm", "RMSE/mm", "中位绝对误差/mm", "最大绝对误差/mm", "符合度/%", "备注"],
    )

    metrics_json = out_dir / "manual_point_metrics.json"
    metrics_json.write_text(
        json.dumps(
            {
                "input_path": str(input_path),
                "reviewed_point_count": len(error_rows),
                "reviewed_image_count": len(image_rows),
                "abs_tol_mm": abs_tol_mm,
                "rel_tol": rel_tol,
                "metrics": metric_rows,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return {
        "point_errors_csv": point_errors_csv,
        "image_errors_csv": image_errors_csv,
        "metrics_csv": metrics_csv,
        "paper_csv": paper_csv,
        "metrics_json": metrics_json,
    }


def run_gui(points_csv: Path, log=None) -> int:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    from PIL import Image, ImageDraw, ImageFont, ImageTk

    if log:
        log(f"run_gui start: {points_csv}")

    class ManualPointReviewApp:
        def __init__(self, root: tk.Tk, csv_path: Path):
            if log:
                log("app init begin")
            self.root = root
            self.csv_path = csv_path
            self.rows = read_csv_rows(csv_path)
            if log:
                log(f"csv loaded rows={len(self.rows)}")
            if not self.rows:
                raise ValueError(f"No rows found in {csv_path}")
            self.image_ids = sorted(
                {str(row["image_record_id"]) for row in self.rows},
                key=lambda value: int(value) if value.isdigit() else value,
            )
            self.image_to_indices: dict[str, list[int]] = defaultdict(list)
            for idx, row in enumerate(self.rows):
                self.image_to_indices[str(row["image_record_id"])].append(idx)
            for indices in self.image_to_indices.values():
                indices.sort(key=lambda idx: int(float(self.rows[idx].get("point_order") or 0)))
            self.image_pos = 0
            self.point_pos = 0
            self.photo = None
            self.current_image_size = (0, 0)
            self.dirty = False
            self.refreshing_tree = False

            self.root.title("Crack Width Manual Review - 裂缝宽度测点级人工复核")
            self.root.geometry("1360x860")
            self.root.protocol("WM_DELETE_WINDOW", self.close)
            self.root.columnconfigure(0, weight=1)
            self.root.rowconfigure(1, weight=1)

            self._build_header(ttk)
            self._build_body(tk, ttk)
            self._build_footer(ttk)
            self.refresh()
            if log:
                log("app init complete")

        def _build_header(self, ttk):
            bar = ttk.Frame(self.root, padding=(10, 8))
            bar.grid(row=0, column=0, sticky="ew")
            bar.columnconfigure(7, weight=1)
            ttk.Button(bar, text="打开CSV", command=self.open_csv).grid(row=0, column=0, padx=4)
            ttk.Button(bar, text="保存CSV", command=self.save_csv).grid(row=0, column=1, padx=4)
            ttk.Button(bar, text="导出Excel", command=self.export_excel).grid(row=0, column=2, padx=4)
            ttk.Button(bar, text="计算指标", command=self.evaluate).grid(row=0, column=3, padx=4)
            ttk.Button(bar, text="上一图", command=self.prev_image).grid(row=0, column=4, padx=4)
            ttk.Button(bar, text="下一图", command=self.next_image).grid(row=0, column=5, padx=4)
            ttk.Button(bar, text="首个未填", command=self.first_unfilled).grid(row=0, column=6, padx=4)
            self.progress_var = tk.StringVar()
            ttk.Label(bar, textvariable=self.progress_var, anchor="e").grid(row=0, column=7, sticky="e", padx=10)

        def _build_body(self, tk, ttk):
            body = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
            body.grid(row=1, column=0, sticky="nsew", padx=10, pady=8)

            canvas_frame = ttk.Frame(body)
            canvas_frame.columnconfigure(0, weight=1)
            canvas_frame.rowconfigure(0, weight=1)
            self.canvas = tk.Canvas(canvas_frame, bg="#202020", highlightthickness=0)
            self.canvas.grid(row=0, column=0, sticky="nsew")
            self.canvas.bind("<Configure>", lambda _event: self.draw_image())
            body.add(canvas_frame, weight=4)

            panel = ttk.Frame(body, padding=(10, 0))
            panel.columnconfigure(0, weight=1)
            body.add(panel, weight=1)

            self.meta_var = tk.StringVar()
            ttk.Label(panel, textvariable=self.meta_var, justify="left", wraplength=360).grid(row=0, column=0, sticky="ew")

            self.tree = ttk.Treeview(panel, columns=("point", "auto", "manual", "status"), show="headings", height=8)
            for col, title, width in [
                ("point", "测点", 64),
                ("auto", "程序/mm", 90),
                ("manual", "人工/mm", 90),
                ("status", "状态", 80),
            ]:
                self.tree.heading(col, text=title)
                self.tree.column(col, width=width, anchor="center")
            self.tree.grid(row=1, column=0, sticky="ew", pady=(12, 8))
            self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

            form = ttk.Frame(panel)
            form.grid(row=2, column=0, sticky="ew")
            form.columnconfigure(1, weight=1)
            self.point_var = tk.StringVar()
            self.auto_var = tk.StringVar()
            self.manual_var = tk.StringVar()
            self.status_var = tk.StringVar()
            self.note_var = tk.StringVar()

            labels = [
                ("当前测点", self.point_var, None),
                ("程序宽度", self.auto_var, None),
                ("人工宽度/mm", self.manual_var, "entry"),
                ("状态", self.status_var, "combo"),
                ("备注", self.note_var, "entry"),
            ]
            for row_idx, (label, variable, kind) in enumerate(labels):
                ttk.Label(form, text=label).grid(row=row_idx, column=0, sticky="w", pady=4)
                if kind == "entry":
                    widget = ttk.Entry(form, textvariable=variable)
                elif kind == "combo":
                    widget = ttk.Combobox(form, textvariable=variable, values=["pending", "reviewed", "skipped"], state="readonly")
                else:
                    widget = ttk.Label(form, textvariable=variable)
                widget.grid(row=row_idx, column=1, sticky="ew", pady=4)

            actions = ttk.Frame(panel)
            actions.grid(row=3, column=0, sticky="ew", pady=12)
            actions.columnconfigure((0, 1), weight=1)
            ttk.Button(actions, text="保存当前", command=self.save_current).grid(row=0, column=0, sticky="ew", padx=3, pady=3)
            ttk.Button(actions, text="跳过测点", command=self.skip_current).grid(row=0, column=1, sticky="ew", padx=3, pady=3)
            ttk.Button(actions, text="上一测点", command=self.prev_point).grid(row=1, column=0, sticky="ew", padx=3, pady=3)
            ttk.Button(actions, text="下一测点", command=self.next_point).grid(row=1, column=1, sticky="ew", padx=3, pady=3)

            self.path_var = tk.StringVar()
            ttk.Label(panel, textvariable=self.path_var, wraplength=360, foreground="#555").grid(row=4, column=0, sticky="ew")

        def _build_footer(self, ttk):
            self.status_line = tk.StringVar()
            ttk.Label(self.root, textvariable=self.status_line, anchor="w", padding=(10, 6)).grid(row=2, column=0, sticky="ew")

        def current_indices(self) -> list[int]:
            return self.image_to_indices[self.image_ids[self.image_pos]]

        def current_row_index(self) -> int:
            indices = self.current_indices()
            self.point_pos = max(0, min(self.point_pos, len(indices) - 1))
            return indices[self.point_pos]

        def current_row(self) -> dict[str, Any]:
            return self.rows[self.current_row_index()]

        def save_current(self) -> bool:
            if not self.commit_current_fields():
                return False
            self.refresh_tree()
            self.update_progress()
            self.status_line.set("当前测点已暂存，点击“保存CSV”写入文件。")
            return True

        def commit_current_fields(self) -> bool:
            row = self.current_row()
            manual = self.manual_var.get().strip()
            if manual:
                try:
                    float(manual)
                except ValueError:
                    messagebox.showerror("输入错误", "人工宽度必须是数字，单位 mm。")
                    return False
            row["manual_width_mm"] = manual
            status = self.status_var.get().strip() or "pending"
            if manual and status == "pending":
                status = "reviewed"
            row["review_status"] = status
            row["manual_note"] = self.note_var.get().strip()
            self.dirty = True
            return True

        def skip_current(self):
            self.status_var.set("skipped")
            if not self.note_var.get().strip():
                self.note_var.set("人工判定该测点不可读")
            if self.save_current():
                self.next_point()

        def save_csv(self):
            if self.save_current():
                write_csv_rows(self.csv_path, self.rows, POINT_FIELDS)
                self.dirty = False
                self.status_line.set(f"已保存：{self.csv_path}")

        def open_csv(self):
            chosen = filedialog.askopenfilename(filetypes=[("CSV", "*.csv")], initialdir=str(self.csv_path.parent))
            if not chosen:
                return
            if self.dirty and not messagebox.askyesno("未保存", "当前 CSV 有未保存修改，仍要打开新文件吗？"):
                return
            for child in self.root.winfo_children():
                child.destroy()
            self.__init__(self.root, Path(chosen))

        def export_excel(self):
            self.save_current()
            xlsx_path = self.csv_path.with_suffix(".xlsx")
            write_csv_rows(self.csv_path, self.rows, POINT_FIELDS)
            create_point_excel(self.csv_path, xlsx_path)
            messagebox.showinfo("导出完成", f"已导出：\n{xlsx_path}")

        def evaluate(self):
            self.save_current()
            write_csv_rows(self.csv_path, self.rows, POINT_FIELDS)
            out_dir = self.csv_path.parent / "manual_point_review_eval"
            outputs = evaluate_points_file(self.csv_path, out_dir, abs_tol_mm=0.02, rel_tol=0.15)
            messagebox.showinfo("计算完成", "\n".join(str(path) for path in outputs.values()))

        def refresh(self):
            self.load_current_fields()
            self.refresh_tree()
            self.draw_image()
            self.update_progress()

        def load_current_fields(self):
            row = self.current_row()
            indices = self.current_indices()
            self.point_var.set(f"{row.get('point_id')}  位置约 {row.get('target_position_pct')}%")
            self.auto_var.set(f"{row.get('auto_width_mm')} mm")
            self.manual_var.set(str(row.get("manual_width_mm") or ""))
            self.status_var.set(str(row.get("review_status") or "pending"))
            self.note_var.set(str(row.get("manual_note") or ""))
            self.meta_var.set(
                f"图像 {self.image_pos + 1}/{len(self.image_ids)}，测点 {self.point_pos + 1}/{len(indices)}\n"
                f"构件：{row.get('device_name')}\n"
                f"时间：{row.get('uptime')}\n"
                f"文件：{row.get('filename')}"
            )
            self.path_var.set(str(row.get("input_image_path") or ""))

        def refresh_tree(self):
            self.refreshing_tree = True
            self.tree.delete(*self.tree.get_children())
            current_idx = self.current_row_index()
            selected_item = None
            for idx in self.current_indices():
                row = self.rows[idx]
                item = self.tree.insert(
                    "",
                    "end",
                    iid=str(idx),
                    values=(
                        f"P{row.get('point_order')}",
                        row.get("auto_width_mm", ""),
                        row.get("manual_width_mm", ""),
                        row.get("review_status", ""),
                    ),
                )
                if idx == current_idx:
                    selected_item = item
            if selected_item:
                self.tree.selection_set(selected_item)
                self.tree.focus(selected_item)
            self.refreshing_tree = False

        def on_tree_select(self, _event):
            if self.refreshing_tree:
                return
            selected = self.tree.selection()
            if not selected:
                return
            idx = int(selected[0])
            indices = self.current_indices()
            if idx in indices:
                if idx == self.current_row_index():
                    return
                if not self.commit_current_fields():
                    self.refresh_tree()
                    return
                self.point_pos = indices.index(idx)
                self.load_current_fields()
                self.draw_image()
                self.refresh_tree()
                self.update_progress()

        def draw_image(self):
            row = self.current_row()
            image_path = Path(str(row.get("input_image_path") or ""))
            self.canvas.delete("all")
            width = max(200, self.canvas.winfo_width())
            height = max(200, self.canvas.winfo_height())
            if not image_path.exists():
                self.canvas.create_text(width // 2, height // 2, text=f"图像不存在：\n{image_path}", fill="white")
                return
            try:
                image = Image.open(image_path).convert("RGB")
            except Exception as exc:
                self.canvas.create_text(width // 2, height // 2, text=f"图像读取失败：\n{exc}", fill="white")
                return
            scale = min(width / image.width, height / image.height)
            display_size = (max(1, int(image.width * scale)), max(1, int(image.height * scale)))
            display = image.resize(display_size)
            draw = ImageDraw.Draw(display)
            font = ImageFont.load_default()
            offset_x = (width - display_size[0]) // 2
            offset_y = (height - display_size[1]) // 2
            current_idx = self.current_row_index()
            for idx in self.current_indices():
                point = self.rows[idx]
                color = "#ffcc00" if idx == current_idx else "#00d1ff"
                line_color = "#ff3b30" if idx == current_idx else "#00a7c8"
                coords = [to_float(point.get(key)) for key in ("left_x_px", "left_y_px", "right_x_px", "right_y_px", "x_px", "y_px")]
                if any(value is None for value in coords):
                    continue
                lx, ly, rx, ry, x, y = [float(value) for value in coords]
                lx, ly, rx, ry, x, y = [value * scale for value in (lx, ly, rx, ry, x, y)]
                draw.line((lx, ly, rx, ry), fill=line_color, width=3 if idx == current_idx else 2)
                radius = 6 if idx == current_idx else 4
                draw.ellipse((x - radius, y - radius, x + radius, y + radius), fill=color, outline="black")
                draw.text((x + 8, y - 12), f"P{point.get('point_order')}", fill=color, font=font)
            self.photo = ImageTk.PhotoImage(display)
            self.canvas.create_image(offset_x, offset_y, anchor="nw", image=self.photo)

        def update_progress(self):
            reviewed = 0
            skipped = 0
            for row in self.rows:
                if is_skipped(row):
                    skipped += 1
                elif to_float(row.get("manual_width_mm")) is not None:
                    reviewed += 1
            total = len(self.rows)
            self.progress_var.set(f"已填 {reviewed}/{total}，跳过 {skipped}，文件：{self.csv_path.name}")

        def move_to(self, image_pos: int, point_pos: int):
            self.save_current()
            self.image_pos = max(0, min(image_pos, len(self.image_ids) - 1))
            self.point_pos = max(0, min(point_pos, len(self.current_indices()) - 1))
            self.refresh()

        def next_point(self):
            if self.point_pos + 1 < len(self.current_indices()):
                self.move_to(self.image_pos, self.point_pos + 1)
            elif self.image_pos + 1 < len(self.image_ids):
                self.move_to(self.image_pos + 1, 0)

        def prev_point(self):
            if self.point_pos > 0:
                self.move_to(self.image_pos, self.point_pos - 1)
            elif self.image_pos > 0:
                prev_indices = self.image_to_indices[self.image_ids[self.image_pos - 1]]
                self.move_to(self.image_pos - 1, len(prev_indices) - 1)

        def next_image(self):
            self.move_to(self.image_pos + 1, 0)

        def prev_image(self):
            self.move_to(self.image_pos - 1, 0)

        def first_unfilled(self):
            self.save_current()
            for image_pos, image_id in enumerate(self.image_ids):
                for point_pos, idx in enumerate(self.image_to_indices[image_id]):
                    row = self.rows[idx]
                    if not is_skipped(row) and to_float(row.get("manual_width_mm")) is None:
                        self.image_pos = image_pos
                        self.point_pos = point_pos
                        self.refresh()
                        return
            messagebox.showinfo("完成", "没有未填写的测点。")

        def close(self):
            if self.dirty:
                result = messagebox.askyesnocancel("未保存", "是否保存 CSV 后退出？")
                if result is None:
                    return
                if result:
                    write_csv_rows(self.csv_path, self.rows, POINT_FIELDS)
            self.root.destroy()

    root = tk.Tk()
    ManualPointReviewApp(root, points_csv)
    root.deiconify()
    root.lift()
    try:
        root.attributes("-topmost", True)
        root.after(1200, lambda: root.attributes("-topmost", False))
    except tk.TclError:
        pass
    root.update()
    if log:
        log("enter mainloop")
    root.mainloop()
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Point-level manual review workflow for crack-width measurements.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    build = subparsers.add_parser("build", help="Build point-level review CSV/XLSX from image summary and profile CSV files.")
    build.add_argument("--summary-csv", type=Path, default=DEFAULT_SUMMARY)
    build.add_argument("--output-csv", type=Path, default=DEFAULT_POINTS_CSV)
    build.add_argument("--output-xlsx", type=Path, default=DEFAULT_POINTS_XLSX)
    build.add_argument("--points-per-image", type=int, default=3)
    build.add_argument("--point-config-csv", type=Path, default=DEFAULT_POINT_CONFIG_CSV)
    build.add_argument("--no-point-config", action="store_true")
    build.add_argument(
        "--excluded-csv",
        type=Path,
        default=Path("outputs/platform_unmarked_full_20260527_v2/manual_point_review_excluded_images.csv"),
    )
    build.add_argument("--allow-marker-points", action="store_true")
    build.add_argument("--no-xlsx", action="store_true")

    init_config = subparsers.add_parser("init-config", help="Create a sensor-level point config CSV from existing point rows.")
    init_config.add_argument("--points-csv", type=Path, default=DEFAULT_POINTS_CSV)
    init_config.add_argument("--output-csv", type=Path, default=DEFAULT_POINT_CONFIG_CSV)
    init_config.add_argument("--search-radius-px", type=float, default=220.0)

    gui = subparsers.add_parser("gui", help="Open the point-level manual review GUI.")
    gui.add_argument("--points-csv", type=Path, default=DEFAULT_POINTS_CSV)

    evaluate = subparsers.add_parser("evaluate", help="Evaluate a filled point-level CSV/XLSX.")
    evaluate.add_argument("--input", type=Path, default=DEFAULT_POINTS_CSV)
    evaluate.add_argument("--out-dir", type=Path, default=DEFAULT_EVAL_DIR)
    evaluate.add_argument("--abs-tol-mm", type=float, default=0.02)
    evaluate.add_argument("--rel-tol", type=float, default=0.15)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    if args.command == "build":
        rows = build_point_rows(
            args.summary_csv,
            args.points_per_image,
            avoid_markers=not args.allow_marker_points,
            excluded_csv=args.excluded_csv,
            point_config_csv=None if args.no_point_config else args.point_config_csv,
        )
        write_csv_rows(args.output_csv, rows, POINT_FIELDS)
        print(args.output_csv)
        print(f"points={len(rows)}")
        print(args.excluded_csv)
        if not args.no_xlsx:
            xlsx = create_point_excel(args.output_csv, args.output_xlsx)
            print(xlsx)
    elif args.command == "init-config":
        rows = read_csv_rows(args.points_csv)
        config_rows = infer_point_config_from_points(rows, search_radius_px=args.search_radius_px)
        write_point_config(args.output_csv, config_rows)
        print(args.output_csv)
        print(f"config_points={len(config_rows)}")
    elif args.command == "gui":
        if not args.points_csv.exists():
            raise FileNotFoundError(f"{args.points_csv} does not exist. Run: python manual_point_review.py build")
        return run_gui(args.points_csv)
    elif args.command == "evaluate":
        outputs = evaluate_points_file(args.input, args.out_dir, args.abs_tol_mm, args.rel_tol)
        for path in outputs.values():
            print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
